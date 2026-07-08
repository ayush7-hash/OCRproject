#!/usr/bin/env python3
"""
screenshot_to_excel_paddleocr_plain.py

Convert a screenshot of a spreadsheet into a clean .xlsx file using
PaddleOCR (via img2table) for table detection and text extraction.

No confidence scoring, no per-cell re-OCR, no highlighting - just a single
OCR pass per image and a straight write-out to Excel. This intentionally
removes the per-cell confidence-check step used in earlier versions of
this script, since that step ran a full separate OCR call on every cell
and was the main cause of long runtimes / high CPU usage.

Requires: pip install paddlepaddle paddleocr img2table openpyxl pillow numpy

Usage:
  python3 screenshot_to_excel_paddleocr_plain.py input.png -o output.xlsx
  python3 screenshot_to_excel_paddleocr_plain.py shot1.png shot2.png -o combined.xlsx
  python3 screenshot_to_excel_paddleocr_plain.py screenshots_folder/ -o output.xlsx
  python3 screenshot_to_excel_paddleocr_plain.py input.png -o output.xlsx --lang en

  for venv --> source venv/bin/activate
"""

import argparse
import json
import sys
import time
from pathlib import Path

from img2table.document import Image
from img2table.ocr import PaddleOCR as I2TPaddleOCR
from openpyxl import Workbook


SUPPORTED_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}


def collect_input_images(paths):
    images = []
    from_directory = False
    for p in paths:
        path = Path(p)
        if not path.exists():
            print(f"WARNING: path does not exist, skipping: {path}", file=sys.stderr)
            continue
        if path.is_dir():
            from_directory = True
            for f in sorted(path.iterdir()):
                if f.suffix.lower() in SUPPORTED_EXTS:
                    images.append(f)
        elif path.suffix.lower() in SUPPORTED_EXTS:
            images.append(path)
        else:
            print(f"WARNING: unsupported file type, skipping: {path}", file=sys.stderr)
    return images, from_directory


BLANK_VALUES = {"", "nan", "none", "null", "na", "n/a", "<na>"}


def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in BLANK_VALUES:
        return None
    return s


def autofit_columns(ws, padding=2, max_width=60):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(width + padding, max_width)


def unique_sheet_name(wb, base):
    name = base or "Sheet"
    i = 1
    existing = set(wb.sheetnames)
    candidate = name
    while candidate in existing:
        i += 1
        candidate = f"{name}_{i}"
    return candidate


def write_table_to_sheet(table, ws):
    for r_idx, cells in enumerate(table.content.values(), start=1):
        for c_idx, cell in enumerate(cells, start=1):
            ws.cell(row=r_idx, column=c_idx, value=clean(cell.value))
    autofit_columns(ws)


def convert(image_paths, output_path, lang="en", min_confidence=50, show_progress=False):
    table_ocr = I2TPaddleOCR(lang=lang)

    wb = Workbook()
    wb.remove(wb.active)

    total = len(image_paths)
    failed = []
    sheet_to_image = {}

    for idx, img_path in enumerate(image_paths, start=1):
        prefix = f"[{idx}/{total}] " if show_progress else ""
        print(f"{prefix}Processing {img_path.name} ...", flush=True)
        t0 = time.time()

        try:
            doc = Image(src=str(img_path), detect_rotation=True)
            tables = doc.extract_tables(
                ocr=table_ocr,
                implicit_rows=True,
                implicit_columns=True,
                borderless_tables=True,
                min_confidence=min_confidence,
            )
        except Exception as e:
            print(f"  ERROR: failed to process {img_path.name} ({e}), skipping.", file=sys.stderr)
            failed.append(img_path)
            continue

        print(f"  Table extraction took {time.time() - t0:.2f}s", flush=True)

        if not tables:
            print(f"  No table detected in {img_path.name}, skipping.", file=sys.stderr)
            continue

        for t_idx, table in enumerate(tables):
            sheet_name = img_path.stem[:25]
            if len(tables) > 1:
                sheet_name += f"_t{t_idx+1}"
            sheet_name = unique_sheet_name(wb, sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            write_table_to_sheet(table, ws)
            sheet_to_image[sheet_name] = str(img_path.resolve())
            print(f"  -> wrote sheet '{sheet_name}' ({table.df.shape[0]} rows x {table.df.shape[1]} cols)")

    if not wb.sheetnames:
        print("ERROR: no tables detected in any input image.", file=sys.stderr)
        sys.exit(1)

    wb.save(output_path)
    print(f"\nSaved: {output_path}")

    mapping_path = Path(output_path).with_suffix(".sheetmap.json")
    with open(mapping_path, "w") as f:
        json.dump(sheet_to_image, f, indent=2)
    print(f"Saved sheet->image mapping: {mapping_path}")

    if failed:
        print(f"\nWARNING: {len(failed)} image(s) failed and were skipped:", file=sys.stderr)
        for f in failed:
            print(f"  - {f}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("inputs", nargs="+", help="Image file(s) or folder(s) of screenshots")
    parser.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    parser.add_argument("--lang", default="en", help="PaddleOCR language code, e.g. en, ch, hi (single language only)")
    parser.add_argument("--min-confidence", type=int, default=50, help="Minimum table-detection confidence 0-100 (default 50)")
    args = parser.parse_args()

    images, from_directory = collect_input_images(args.inputs)
    if not images:
        print("ERROR: no valid image files found.", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    show_progress = from_directory and len(images) > 1

    convert(images, output_path, lang=args.lang, min_confidence=args.min_confidence, show_progress=show_progress)


if __name__ == "__main__":
    main()
