#!/usr/bin/env python3
"""
verify_with_gemini.py

Second pass after screenshot_to_excel_paddleocr_plain.py. Takes the .xlsx
produced by that script (plus its sidecar .sheetmap.json), and for each
sheet sends the original screenshot + the current OCR'd cell grid to
Gemini (vision) to cross-check every cell against the image.

The FINAL output of this script is a single JSON file containing the
corrected data for every sheet. The intermediate .xlsx (from the OCR step)
and its .sheetmap.json are deleted afterward by default, so the only
artifact left on disk is the corrected JSON.

Requires: pip install google-genai openpyxl
Requires environment variable: GEMINI_API_KEY

Usage:
  python3 verify_with_gemini.py output.xlsx
  python3 verify_with_gemini.py output.xlsx --model gemini-2.5-flash
  python3 verify_with_gemini.py output.xlsx --sheetmap output.sheetmap.json
  python3 verify_with_gemini.py output.xlsx --output final_result.json
  python3 verify_with_gemini.py output.xlsx --keep-intermediate   # don't delete xlsx/sheetmap
"""

import argparse
import json
import mimetypes
import sys
import time
from pathlib import Path

from google import genai
from google.genai import types
from google.genai.errors import APIError
from openpyxl import load_workbook

DEFAULT_MODEL = "gemini-2.5-flash"
MAX_TOKENS = 64000
MAX_RETRIES = 5
RETRY_BASE_DELAY = 5  # seconds, doubles each retry


def call_with_retry(fn, *args, **kwargs):
    """Call fn, retrying on transient errors (429 rate limit, 503 overloaded)."""
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except APIError as e:
            status = getattr(e, "code", None)
            if status not in (429, 503) or attempt == MAX_RETRIES:
                raise
            delay = RETRY_BASE_DELAY * (2 ** (attempt - 1))
            print(
                f"  {status} error, retrying in {delay}s (attempt {attempt}/{MAX_RETRIES}) ...",
                file=sys.stderr,
            )
            time.sleep(delay)
            last_err = e
    raise last_err


def resolve_sheetmap_path(xlsx_path, sheetmap_path):
    if sheetmap_path is None:
        return Path(xlsx_path).with_suffix(".sheetmap.json")
    return Path(sheetmap_path)


def load_sheetmap(xlsx_path, sheetmap_path):
    sheetmap_path = resolve_sheetmap_path(xlsx_path, sheetmap_path)
    if not sheetmap_path.exists():
        print(
            f"ERROR: sheet map not found at {sheetmap_path}. "
            "Run the OCR script first (it now writes this file automatically), "
            "or pass --sheetmap explicitly.",
            file=sys.stderr,
        )
        sys.exit(1)
    with open(sheetmap_path) as f:
        return json.load(f)


def sheet_to_grid(ws):
    grid = []
    for row in ws.iter_rows():
        grid.append([cell.value for cell in row])
    return grid


def workbook_to_json(wb):
    data = {}
    for sheet_name in wb.sheetnames:
        data[sheet_name] = sheet_to_grid(wb[sheet_name])
    return data


def read_image_bytes(img_path):
    media_type = mimetypes.guess_type(str(img_path))[0] or "image/png"
    with open(img_path, "rb") as f:
        data = f.read()
    return media_type, data


def build_prompt(grid):
    # Represent the current grid as simple pipe-delimited rows so the model
    # can reference cells positionally without any ambiguity.
    lines = []
    for r_idx, row in enumerate(grid):
        cells = ["" if v is None else str(v) for v in row]
        lines.append(f"Row {r_idx}: " + " | ".join(cells))
    grid_text = "\n".join(lines)

    return f"""You are cross-checking OCR-extracted spreadsheet data against the original screenshot image.

The image shows a table. Below is the data currently extracted via OCR, given as one line per row,
with cells separated by " | " (in the same left-to-right order as they appear in the image):

{grid_text}

Compare every cell against the image carefully. OCR errors typically look like:
- digit misreads (0/O, 1/l/I, 5/S, 8/B)
- misplaced or dropped decimal points / commas
- merged or split words
- extra/missing whitespace
- wrong case

Only fix cells where the image clearly shows something different from the extracted text.
Do NOT invent, guess, or "improve" data that isn't visibly wrong in the image.
CRITICAL: Every single output row MUST have exactly {len(grid[0]) if grid else 0} cells, even if
some of those cells are empty strings "". Never merge two cells into one, and never omit a cell,
even if it looks blank or repeated in the image.
Empty cells should be represented as an empty string "".

Respond with ONLY a JSON object, no other text, no markdown fences, in this exact shape:
{{"rows": [["cell", "cell", ...], ["cell", "cell", ...], ...]}}
"""


def parse_response(text, expected_rows, expected_cols):
    text = text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ValueError(f"model did not return valid JSON: {e}")

    rows = data.get("rows")
    if not isinstance(rows, list):
        raise ValueError("JSON missing a 'rows' list")
    if len(rows) != expected_rows:
        raise ValueError(f"row count mismatch: expected {expected_rows}, got {len(rows)}")

    fixed_rows = []
    mismatches = []
    for r_idx, r in enumerate(rows):
        if len(r) != expected_cols:
            mismatches.append((r_idx, len(r)))
            if len(r) < expected_cols:
                r = r + [""] * (expected_cols - len(r))
            else:
                r = r[:expected_cols]
        fixed_rows.append(r)

    return fixed_rows, mismatches


def verify_sheet(client, model, ws, img_path):
    grid = sheet_to_grid(ws)
    if not grid:
        return None, "empty sheet, skipped"

    n_rows = len(grid)
    n_cols = len(grid[0])

    media_type, img_bytes = read_image_bytes(img_path)
    prompt = build_prompt(grid)

    response = call_with_retry(
        client.models.generate_content,
        model=model,
        contents=[
            types.Part.from_bytes(data=img_bytes, mime_type=media_type),
            types.Part.from_text(text=prompt),
        ],
        config=types.GenerateContentConfig(
            max_output_tokens=MAX_TOKENS,
            response_mime_type="application/json",
        ),
    )

    full_text = response.text or ""

    finish_reason = None
    if response.candidates:
        finish_reason = response.candidates[0].finish_reason

    if finish_reason == "MAX_TOKENS":
        return None, (
            f"response was truncated (hit MAX_TOKENS={MAX_TOKENS}); "
            "the sheet is likely too large for one request, increase MAX_TOKENS "
            "further or split the sheet into smaller regions"
        )

    if not full_text.strip():
        return None, f"empty response from model (finish_reason={finish_reason})"

    try:
        corrected_rows, mismatches = parse_response(full_text, n_rows, n_cols)
    except ValueError as e:
        return None, f"could not apply correction ({e}); finish_reason={finish_reason}"

    if mismatches:
        details = ", ".join(f"row {r} (got {n} cols)" for r, n in mismatches)
        print(
            f"  NOTE: {len(mismatches)} row(s) had a column-count mismatch and were "
            f"padded/trimmed to fit: {details}. Worth spot-checking those rows.",
            file=sys.stderr,
        )

    changes = 0
    for r_idx, row in enumerate(corrected_rows):
        for c_idx, new_val in enumerate(row):
            new_val = new_val if new_val != "" else None
            old_val = grid[r_idx][c_idx]
            if new_val != old_val:
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=new_val)
                changes += 1

    return changes, None


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx", help="Path to the .xlsx produced by the OCR script")
    parser.add_argument("--sheetmap", default=None, help="Path to sidecar .sheetmap.json (default: inferred from xlsx name)")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"Gemini model to use (default: {DEFAULT_MODEL})")
    parser.add_argument("--output", default=None, help="Path for the final corrected JSON (default: <xlsx name>.final.json)")
    parser.add_argument(
        "--keep-intermediate",
        action="store_true",
        help="Keep the intermediate .xlsx and .sheetmap.json instead of deleting them after the final JSON is written",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    sheetmap_path = resolve_sheetmap_path(xlsx_path, args.sheetmap)
    sheetmap = load_sheetmap(xlsx_path, args.sheetmap)
    client = genai.Client()  # reads GEMINI_API_KEY from env

    wb = load_workbook(xlsx_path)

    total = len(wb.sheetnames)
    failed = []
    total_changes = 0

    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        print(f"[{idx}/{total}] Verifying sheet '{sheet_name}' ...", flush=True)

        img_path = sheetmap.get(sheet_name)
        if not img_path or not Path(img_path).exists():
            print(f"  WARNING: no source image found for sheet '{sheet_name}', skipping.", file=sys.stderr)
            failed.append(sheet_name)
            continue

        ws = wb[sheet_name]
        try:
            changes, err = verify_sheet(client, args.model, ws, img_path)
        except Exception as e:
            print(f"  ERROR: {e}", file=sys.stderr)
            failed.append(sheet_name)
            continue

        if err:
            print(f"  WARNING: {err}, sheet left unchanged.", file=sys.stderr)
            failed.append(sheet_name)
            continue

        print(f"  -> {changes} cell(s) corrected")
        total_changes += changes

    # Corrections above were applied to `wb` in memory (ws.cell(...) calls).
    # Export that corrected state straight to JSON rather than re-saving the xlsx.
    output_json_path = Path(args.output) if args.output else xlsx_path.with_suffix(".final.json")
    final_data = workbook_to_json(wb)
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(final_data, f, indent=2, ensure_ascii=False)

    print(f"\nSaved final corrected data as JSON: {output_json_path}")
    print(f"Total cells corrected: {total_changes}")

    if failed:
        print(f"\nWARNING: {len(failed)} sheet(s) could not be verified:", file=sys.stderr)
        for s in failed:
            print(f"  - {s}", file=sys.stderr)

    if args.keep_intermediate:
        print(f"\nKeeping intermediate files (--keep-intermediate set): {xlsx_path}, {sheetmap_path}")
    else:
        for p in (xlsx_path, sheetmap_path):
            try:
                Path(p).unlink()
                print(f"Removed intermediate file: {p}")
            except FileNotFoundError:
                pass
            except OSError as e:
                print(f"WARNING: could not remove {p}: {e}", file=sys.stderr)
        print(f"\nOutput File: {output_json_path}")


if __name__ == "__main__":
    main()
